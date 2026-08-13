"""
=====================================================================
 B.Sc. COUNSELLING PORTAL — SESSION 2026-27  (Botany interface)
 Aligarh Muslim University, Aligarh
---------------------------------------------------------------------
 Flask + SQLite web application (self-contained; no external services)

 Modules:
   1. Counselling portal  — registrations, mentor appointments,
      grievances, notices, admin dashboard.
   2. Counselling Data Entry ("Botany Counselling" interface):
      Main Entry Sheet (one row per student, searchable course-code
      cells for VAC V / VOC V / VAC VI / VOC VI), automatic course
      separation into per-course sheets, Excel/CSV import & export,
      blank Excel format + sample file, PDF export, ZIP export,
      printables, database backup/restore.            [Admin only]

 Run     :  python3 app.py
 Portal  :  http://localhost:5000
 Admin   :  http://localhost:5000/admin   (default password: amu@2026)

 Change the admin password with:  ADMIN_PASSWORD=xxxx python3 app.py
=====================================================================
"""

import csv
import io
import os
import re
import shutil
import sqlite3
import tempfile
import zipfile
from datetime import date, datetime
from functools import wraps

from flask import (Flask, Response, abort, flash, g, jsonify, redirect,
                   render_template, request, send_file, session, url_for)

# ---------------------------------------------------------------- config
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.environ.get("DATABASE_PATH",
                         os.path.join(BASE_DIR, "counselling.db"))

ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "amu@2026")
SECRET_KEY = os.environ.get("SECRET_KEY", "bsc-counselling-2026-27-secret")
PORT = int(os.environ.get("PORT", "5000"))

try:                                       # PDF export (optional extra)
    import reportlab                       # noqa: F401
    HAS_PDF = True
except Exception:
    HAS_PDF = False

try:                                       # QR codes on student rows
    import qrcode                          # noqa: F401
    HAS_QR = True
except Exception:
    HAS_QR = False

app = Flask(__name__)
app.secret_key = SECRET_KEY
# trust reverse-proxy headers (Render/Railway/Nginx) when one is in front
from werkzeug.middleware.proxy_fix import ProxyFix   # noqa: E402
if os.environ.get("BEHIND_PROXY", "1") == "1":
    app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1)
# ---- session/cookie hardening (set FLASK_SECURE_COOKIES=1 behind HTTPS)
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
    SESSION_COOKIE_SECURE=os.environ.get("FLASK_SECURE_COOKIES", "0") == "1",
    MAX_CONTENT_LENGTH=int(os.environ.get("MAX_UPLOAD_MB", "10")) * 1024
    * 1024,
)

# ------------------------------------------------------------------ data
TIME_SLOTS = [
    "10:00 AM – 10:30 AM",
    "10:30 AM – 11:00 AM",
    "11:00 AM – 11:30 AM",
    "11:30 AM – 12:00 PM",
    "12:00 PM – 12:30 PM",
    "02:00 PM – 02:30 PM",
    "02:30 PM – 03:00 PM",
    "03:00 PM – 03:30 PM",
    "03:30 PM – 04:00 PM",
]

SUBJECT_COMBOS = [
    "Physics, Chemistry, Mathematics (PCM)",
    "Physics, Mathematics, Computer Science",
    "Chemistry, Botany, Zoology (CBZ)",
    "Chemistry, Botany, Geology",
    "Mathematics, Statistics, Computer Science",
    "Botany, Zoology, Chemistry",
]

CATEGORIES = ["General", "OBC (NCL)", "SC", "ST", "EWS", "PwD", "Other"]

GRIEVANCE_CATEGORIES = [
    "Academic / Mentoring",
    "Examination / Evaluation",
    "Admission / Allotment",
    "Administrative",
    "Scholarship / Financial",
    "Hostel / Campus",
    "Other",
]

# ---------------------------------------------------- counselling module
from oe_catalog import OE_CATALOG          # parsed official V-sem code list

OEC_SEMS = ["V", "VI"]
OEC_TYPES = ["VAC", "VOC"]
OEC_TYPES_ALL = ["VAC", "VOC", "MO", "NO", "XO", "MJ"]

# slot key, course type, semester, label — drives tabs, columns, validations
OEC_SLOTS = [
    ("vac_v",  "VAC", "V",  "VAC V"),
    ("voc_v",  "VOC", "V",  "VOC V"),
    ("mo_v",   "MO",  "V",  "MO V"),
    ("no_v",   "NO",  "V",  "NO V"),
    ("xo_v",   "XO",  "V",  "XO V"),
    ("vac_vi", "VAC", "VI", "VAC VI"),
    ("voc_vi", "VOC", "VI", "VOC VI"),
    ("mo_vi",  "MO",  "VI", "MO VI"),
    ("no_vi",  "NO",  "VI", "NO VI"),
    ("xo_vi",  "XO",  "VI", "XO VI"),
]
OEC_SLOT_MAP = {s[0]: s for s in OEC_SLOTS}

OEC_STATUSES = ["Pending", "Counselling Done", "Seat Allotted",
                "Reported", "Cancelled"]

# columns of the Main Entry Sheet (Excel / print layout) — generated from slots
_STU_BASE_KEYS = ["enrolment_no", "student_name", "faculty_no", "mobile",
                  "semester"]
_STU_TAIL_KEYS = ["counselling_date", "status", "remarks"]
STU_KEYS = _STU_BASE_KEYS + [s[0] for s in OEC_SLOTS] + _STU_TAIL_KEYS
STU_HEADERS = (["Enrollment No", "Student Name", "Faculty No.", "Mobile",
                "Semester"]
               + ["%s (Course Code)" % s[3] for s in OEC_SLOTS]
               + ["Counselling Date", "Status", "Remarks"])
STU_SHEET_HEADERS = ["S.No"] + STU_HEADERS
STU_WIDTHS = ([6, 15, 26, 14, 14, 9] + [27] * len(OEC_SLOTS)
              + [14, 13, 18])

SPLIT_HEADERS = ["S.No", "Enrollment No", "Student Name", "Faculty No.",
                 "Mobile", "Course Code", "Course Title",
                 "Counselling Date", "Status", "Remarks"]
SPLIT_KEYS = ["enrolment_no", "student_name", "faculty_no", "mobile",
              "course_code", "course_title", "counselling_date",
              "status", "remarks"]
SPLIT_WIDTHS = [6, 15, 26, 14, 14, 14, 40, 14, 13, 18]

SCHEMA = """
CREATE TABLE IF NOT EXISTS mentors (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    name        TEXT NOT NULL,
    designation TEXT NOT NULL,
    department  TEXT NOT NULL,
    email       TEXT,
    room        TEXT,
    day         TEXT,
    slot        TEXT,
    focus       TEXT,
    created_at  TEXT NOT NULL
);
CREATE TABLE IF NOT EXISTS notices (
    id           INTEGER PRIMARY KEY AUTOINCREMENT,
    title        TEXT NOT NULL,
    body         TEXT NOT NULL,
    tag          TEXT DEFAULT 'General',
    published_on TEXT NOT NULL,
    created_at   TEXT NOT NULL
);
CREATE TABLE IF NOT EXISTS registrations (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    kind        TEXT NOT NULL,
    name        TEXT NOT NULL,
    father_name TEXT,
    dob         TEXT,
    email       TEXT NOT NULL,
    phone       TEXT NOT NULL,
    roll_no     TEXT,
    category    TEXT,
    address     TEXT,
    combo       TEXT,
    pref1       TEXT,
    pref2       TEXT,
    pref3       TEXT,
    percentage  TEXT,
    message     TEXT,
    created_at  TEXT NOT NULL
);
CREATE TABLE IF NOT EXISTS appointments (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    name        TEXT NOT NULL,
    roll_no     TEXT NOT NULL,
    email       TEXT NOT NULL,
    phone       TEXT,
    mentor_id   INTEGER,
    mentor_name TEXT,
    date        TEXT NOT NULL,
    slot        TEXT NOT NULL,
    topic       TEXT,
    status      TEXT DEFAULT 'Pending',
    created_at  TEXT NOT NULL
);
CREATE TABLE IF NOT EXISTS grievances (
    id         INTEGER PRIMARY KEY AUTOINCREMENT,
    name       TEXT NOT NULL,
    roll_no    TEXT,
    email      TEXT NOT NULL,
    category   TEXT,
    subject    TEXT NOT NULL,
    message    TEXT NOT NULL,
    created_at TEXT NOT NULL
);
CREATE TABLE IF NOT EXISTS oec_courses (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    semester    TEXT NOT NULL,
    course_type TEXT NOT NULL,
    course_code TEXT,
    title       TEXT,
    created_at  TEXT NOT NULL
);
CREATE TABLE IF NOT EXISTS oec_students (
    id               INTEGER PRIMARY KEY AUTOINCREMENT,
    enrolment_no     TEXT,
    student_name     TEXT NOT NULL,
    faculty_no       TEXT,
    mobile           TEXT,
    semester         TEXT,
    vac_v_code       TEXT, vac_v_title  TEXT,
    voc_v_code       TEXT, voc_v_title  TEXT,
    vac_vi_code      TEXT, vac_vi_title TEXT,
    voc_vi_code      TEXT, voc_vi_title TEXT,
    mo_v_code        TEXT, mo_v_title   TEXT,
    no_v_code        TEXT, no_v_title   TEXT,
    xo_v_code        TEXT, xo_v_title   TEXT,
    counselling_date TEXT,
    status           TEXT DEFAULT 'Pending',
    remarks          TEXT,
    created_at       TEXT NOT NULL,
    updated_at       TEXT NOT NULL
);
"""

# NOTE: The names below are illustrative placeholders for the template.
SEED_MENTORS = [
    ("Dr. Ayesha Rahman",       "Associate Professor",  "Physics",
     "arahman.sc@amu.ac.in",    "Room 12, Physics Dept.",      "Monday",
     "11:00 AM – 12:00 PM",     "Study planning & backlog guidance"),
    ("Prof. S. K. Verma",       "Professor",            "Chemistry",
     "skverma.ch@amu.ac.in",    "Room 5, Chemistry Dept.",     "Tuesday",
     "10:00 AM – 11:00 AM",     "Career guidance (M.Sc. / research)"),
    ("Dr. M. Faizan Qureshi",   "Assistant Professor",  "Mathematics",
     "mfqureshi.mt@amu.ac.in",  "Room 21, Mathematics Dept.",  "Wednesday",
     "02:00 PM – 03:00 PM",     "Exam strategy & time management"),
    ("Dr. Neha Chaturvedi",     "Assistant Professor",  "Botany",
     "nchaturvedi.bt@amu.ac.in","Room 8, Botany Dept.",        "Thursday",
     "11:00 AM – 12:00 PM",     "Academic stress & well-being"),
    ("Prof. R. C. Sharma",      "Professor",            "Zoology",
     "rcsharma.zo@amu.ac.in",   "Room 3, Zoology Dept.",       "Friday",
     "03:00 PM – 04:00 PM",     "Competitive exams (CSIR-UGC NET, GATE)"),
    ("Dr. Imran Ali",           "Assistant Professor",  "Computer Science",
     "iali.cs@amu.ac.in",       "Room 17, Computer Science",   "Monday",
     "02:00 PM – 03:00 PM",     "Programming help & project guidance"),
    ("Dr. Sana Parveen",        "Assistant Professor",  "Statistics",
     "sparveen.st@amu.ac.in",   "Room 9, Statistics Dept.",    "Tuesday",
     "12:00 PM – 01:00 PM",     "Data skills & internship guidance"),
    ("Dr. Arvind Kumar",        "Associate Professor",  "Geology",
     "akumar.gl@amu.ac.in",     "Room 6, Geology Dept.",       "Thursday",
     "03:00 PM – 04:00 PM",     "Field-work prep & higher studies"),
]

SEED_NOTICES = [
    ("Mentor–Mentee registration for B.Sc. V Semester (Session 2026-27) opens on 17 August 2026",
     "All students of B.Sc. V Semester are required to register online for the Mentor–Mentee "
     "scheme through this portal between 17 and 25 August 2026. Each student will be allotted "
     "a faculty mentor from the concerned subject. Students must meet their allotted mentor "
     "at least once every fortnight. Registration is free of cost.", "Academic", "2026-08-17"),
    ("B.Sc. Admission Counselling 2026-27: Round-1 registration and choice filling",
     "Candidates provisionally shortlisted for admission to B.Sc. (Hons.) programmes must complete "
     "counselling registration and fill subject-combination preferences from 18 to 26 August 2026. "
     "Choices once locked cannot be modified. Keep all original documents ready before reporting.",
     "Admission", "2026-08-15"),
    ("Provisional seat allotment list (Round 1) will be displayed on 28 August 2026",
     "The Round-1 seat allotment result will be published on this portal and the University notice "
     "board on 28 August 2026 after 5:00 PM. Allotted candidates must report for document "
     "verification between 29 August and 02 September 2026, failing which the seat will be cancelled.",
     "Admission", "2026-08-12"),
    ("Counselling Cell timings and helpdesk",
     "The Counselling Cell, Faculty of Science, will remain open on all working days from "
     "10:00 AM to 4:30 PM (Friday prayer break 12:30 PM – 2:00 PM). For urgent queries, "
     "students may also write to the helpdesk e-mail given on the Contact page.",
     "General", "2026-08-10"),
    ("First mentor–mentee meeting to be completed by 31 August 2026",
     "All faculty mentors are requested to hold their first structured meeting with allotted "
     "mentees of B.Sc. V Semester by 31 August 2026 and submit the meeting record to the "
     "Counselling Cell by 02 September 2026.", "Academic", "2026-08-08"),
]

# Open-elective course master. The catalogue of official Semester-V codes
# (MO / NO / XO / MJ / VA→VAC / VO→VOC) is parsed from oe_catalog.py; the two
# official VAC courses keep their known titles.
SEED_OEC_TITLES = {
    "WLBLVA5001": "Fundamentals of Environment",
    "WLBLVA6002": "Environmental Education and Conservation",
}
LEGACY_DEMO_CODES = {   # illustrative seeds from the first version — removed
    "GGBSVA5002", "STBSVA5003", "BTBLVA5004", "ZYBLVO5001", "ENBAVO5002",
    "ECBAVO5003", "ENBAVA6004", "GGBSVA6005", "BTBLVA6006", "STBSVO6001",
    "ZYBLVO6003", "ENBAVO6005",
    "WLBLVA5001", "WLBLVA6002",
}

# ---------------------------------------------------------------- helpers
def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
    return g.db


@app.teardown_appcontext
def close_db(exc=None):
    db = g.pop("db", None)
    if db is not None:
        db.close()


def init_db():
    db = sqlite3.connect(DB_PATH)
    db.executescript(SCHEMA)
    # ---- migration path for databases created by earlier versions ----
    db.execute("DROP TABLE IF EXISTS oec_entries")
    have_cols = {r[1] for r in db.execute("PRAGMA table_info(oec_students)")}
    for key in OEC_SLOT_MAP:
        for suffix in ("_code", "_title"):
            col = key + suffix
            if col not in have_cols:
                db.execute("ALTER TABLE oec_students ADD COLUMN %s TEXT" % col)
    have_codes = {r[0] for r in db.execute("SELECT course_code FROM oec_courses")}
    legacy_only = have_codes <= LEGACY_DEMO_CODES if have_codes else True
    if legacy_only:
        db.execute("DELETE FROM oec_courses")
    # -----------------------------------------------------------------
    now = datetime.now().isoformat(timespec="seconds")
    if db.execute("SELECT COUNT(*) FROM mentors").fetchone()[0] == 0:
        db.executemany(
            "INSERT INTO mentors (name, designation, department, email, room,"
            " day, slot, focus, created_at) VALUES (?,?,?,?,?,?,?,?,?)",
            [(*m, now) for m in SEED_MENTORS],
        )
    if db.execute("SELECT COUNT(*) FROM notices").fetchone()[0] == 0:
        db.executemany(
            "INSERT INTO notices (title, body, tag, published_on, created_at)"
            " VALUES (?,?,?,?,?)",
            [(t, b, tag, pub, now) for (t, b, tag, pub) in SEED_NOTICES],
        )
    _seed_catalog(db, now)          # idempotent upsert of the catalogue
    db.commit()
    db.close()


def reseed_courses(db):
    return db.execute("SELECT COUNT(*) FROM oec_courses").fetchone()[0] == 0


def _seed_catalog(db, now):
    """Idempotent upsert of the official V & VI catalogue into the master."""
    from oe_catalog import VOC_VI_OFFERED
    have = {r[0] for r in db.execute("SELECT course_code FROM oec_courses")}
    add = []
    for code, cat, sem in [("WLBLVA5001", "VAC", "V"),
                           ("WLBLVA6002", "VAC", "VI")] + \
                          [(c, t, s) for c, t, s in OE_CATALOG]:
        # VOC VI slot carries only the department's offered list
        if sem == "VI" and cat == "VOC" and code not in VOC_VI_OFFERED:
            used = db.execute("SELECT 1 FROM oec_students WHERE"
                              " voc_vi_code=? LIMIT 1", (code,)).fetchone()
            if not used:
                db.execute("DELETE FROM oec_courses WHERE course_code=?",
                           (code,))
            continue
        if code in have:
            if code in SEED_OEC_TITLES:
                db.execute("UPDATE oec_courses SET title=? WHERE"
                           " course_code=? AND (title='' OR title IS NULL)",
                           (SEED_OEC_TITLES[code], code))
            continue
        have.add(code)
        add.append((sem, cat, code, SEED_OEC_TITLES.get(code, ""), now))
    db.executemany(
        "INSERT INTO oec_courses (semester, course_type, course_code,"
        " title, created_at) VALUES (?,?,?,?,?)", add)


init_db()


def clean(value):
    return (value or "").strip()


@app.context_processor
def inject_globals():
    return {
        "COLLEGE": "Aligarh Muslim University, Aligarh",
        "PORTAL_TITLE": "B.Sc. Counselling Portal",
        "SESSION": "2026-27",
        "BRAND": "Botany Counselling",
        "BRAND_SUB": "B.Sc. Student Portal • 2026-27",
        "DEPT": "Department of Botany",
        "today": date.today().isoformat(),
        "TIME_SLOTS": TIME_SLOTS,
        "SUBJECT_COMBOS": SUBJECT_COMBOS,
        "CATEGORIES": CATEGORIES,
        "GRIEVANCE_CATEGORIES": GRIEVANCE_CATEGORIES,
        "OEC_SEMS": OEC_SEMS,
        "OEC_TYPES": OEC_TYPES,
        "OEC_TYPES_ALL": OEC_TYPES_ALL,
        "OEC_SLOTS": OEC_SLOTS,
        "OEC_STATUSES": OEC_STATUSES,
        "HAS_PDF": HAS_PDF,
        "HAS_QR": HAS_QR,
    }


# ------------------------------------------------------------------ pages
@app.route("/")
def home():
    db = get_db()
    notices = db.execute(
        "SELECT * FROM notices ORDER BY published_on DESC, id DESC LIMIT 5"
    ).fetchall()
    mentor_count = db.execute("SELECT COUNT(*) c FROM mentors").fetchone()["c"]
    return render_template("index.html", notices=notices,
                           mentor_count=mentor_count, active="home")


@app.route("/academic")
def academic():
    return render_template("academic.html", active="academic")


@app.route("/admission")
def admission():
    return render_template("admission.html", active="admission")


@app.route("/mentors")
def mentors():
    db = get_db()
    rows = db.execute("SELECT * FROM mentors ORDER BY department, name").fetchall()
    return render_template("mentors.html", mentors=rows, active="mentors")


@app.route("/schedule")
def schedule():
    db = get_db()
    rows = db.execute("SELECT * FROM mentors ORDER BY department, name").fetchall()
    return render_template("schedule.html", mentors=rows, active="schedule")


@app.route("/notices")
def notices():
    db = get_db()
    rows = db.execute(
        "SELECT * FROM notices ORDER BY published_on DESC, id DESC"
    ).fetchall()
    return render_template("notices.html", notices=rows, active="notices")


@app.route("/faq")
def faq():
    return render_template("faq.html", active="faq")


@app.route("/contact")
def contact():
    return render_template("contact.html", active="contact")


# ------------------------------------------------------------- registration
@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        f = request.form
        kind = clean(f.get("kind")) or "academic"
        if kind not in ("academic", "admission"):
            kind = "academic"

        errors = []
        name = clean(f.get("name"))
        email = clean(f.get("email"))
        phone = clean(f.get("phone"))

        if not name:
            errors.append("Full name is required.")
        if not email or "@" not in email or "." not in email:
            errors.append("A valid e-mail address is required.")
        digits = "".join(ch for ch in phone if ch.isdigit())
        if len(digits) < 10:
            errors.append("A valid 10-digit mobile number is required.")
        if kind == "academic" and not clean(f.get("roll_no")):
            errors.append("Roll / Enrolment number is required for academic "
                          "counselling registration.")
        if kind == "admission" and not clean(f.get("pref1")):
            errors.append("Please select at least Preference-1 of subject "
                          "combination.")

        if errors:
            for e in errors:
                flash(e, "error")
            return render_template("register.html", form=f, active="register"), 400

        db = get_db()
        cur = db.execute(
            """INSERT INTO registrations
               (kind, name, father_name, dob, email, phone, roll_no, category,
                address, combo, pref1, pref2, pref3, percentage, message,
                created_at)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (kind, name, clean(f.get("father_name")), clean(f.get("dob")),
             email, phone, clean(f.get("roll_no")), clean(f.get("category")),
             clean(f.get("address")), clean(f.get("combo")),
             clean(f.get("pref1")), clean(f.get("pref2")), clean(f.get("pref3")),
             clean(f.get("percentage")), clean(f.get("message")),
             datetime.now().isoformat(timespec="seconds")),
        )
        db.commit()
        ref = "R%d-%s" % (cur.lastrowid, "AC" if kind == "academic" else "AD")
        flash("Registration submitted successfully. Your reference ID is "
              "<strong>%s</strong> — please note it down / take a screenshot. "
              "The Counselling Cell will contact you on your e-mail / mobile."
              % ref, "success")
        return redirect(url_for("register"))

    return render_template("register.html", form={}, active="register")


# ------------------------------------------------------------- appointment
@app.route("/appointment", methods=["GET", "POST"])
def appointment():
    db = get_db()
    mentors = db.execute("SELECT * FROM mentors ORDER BY department, name").fetchall()

    if request.method == "POST":
        f = request.form
        errors = []
        name = clean(f.get("name"))
        roll = clean(f.get("roll_no"))
        email = clean(f.get("email"))
        mentor_id = clean(f.get("mentor_id"))
        appt_date = clean(f.get("date"))
        slot = clean(f.get("slot"))
        topic = clean(f.get("topic"))

        if not name:
            errors.append("Name is required.")
        if not roll:
            errors.append("Roll / Enrolment number is required.")
        if not email or "@" not in email:
            errors.append("A valid e-mail address is required.")
        mentor = None
        if mentor_id.isdigit():
            mentor = db.execute("SELECT * FROM mentors WHERE id = ?",
                                (int(mentor_id),)).fetchone()
        if mentor is None:
            errors.append("Please select a mentor.")
        if not appt_date:
            errors.append("Please choose a date.")
        elif appt_date < date.today().isoformat():
            errors.append("Appointment date cannot be in the past.")
        if slot not in TIME_SLOTS:
            errors.append("Please choose a time slot.")

        if not errors and mentor:
            clash = db.execute(
                "SELECT COUNT(*) c FROM appointments WHERE mentor_id = ? AND"
                " date = ? AND slot = ? AND status != 'Cancelled'",
                (mentor["id"], appt_date, slot)).fetchone()["c"]
            if clash:
                errors.append(
                    "That slot is already booked for %s on %s. "
                    "Please choose another time." % (mentor["name"], appt_date))

        if errors:
            for e in errors:
                flash(e, "error")
            return render_template("appointment.html", mentors=mentors,
                                   form=f, active="appointment"), 400

        cur = db.execute(
            """INSERT INTO appointments
               (name, roll_no, email, phone, mentor_id, mentor_name, date,
                slot, topic, status, created_at)
               VALUES (?,?,?,?,?,?,?,?,?,'Pending',?)""",
            (name, roll, email, clean(f.get("phone")), mentor["id"],
             mentor["name"], appt_date, slot, topic,
             datetime.now().isoformat(timespec="seconds")),
        )
        db.commit()
        flash("Appointment request submitted. Your appointment ID is "
              "<strong>A%d</strong> (%s · %s · %s). Please report 10 minutes "
              "early at the mentor's room with your University ID card."
              % (cur.lastrowid, mentor["name"], appt_date, slot), "success")
        return redirect(url_for("appointment"))

    return render_template("appointment.html", mentors=mentors,
                           form={}, active="appointment")


# --------------------------------------------------------------- grievance
@app.route("/grievance", methods=["GET", "POST"])
def grievance():
    if request.method == "POST":
        f = request.form
        errors = []
        name = clean(f.get("name"))
        email = clean(f.get("email"))
        subject = clean(f.get("subject"))
        message = clean(f.get("message"))

        if not name:
            errors.append("Name is required.")
        if not email or "@" not in email:
            errors.append("A valid e-mail address is required.")
        if not subject:
            errors.append("Subject is required.")
        if len(message) < 20:
            errors.append("Please describe the grievance in at least "
                          "20 characters.")

        if errors:
            for e in errors:
                flash(e, "error")
            return render_template("grievance.html", form=f,
                                   active="grievance"), 400

        db = get_db()
        cur = db.execute(
            """INSERT INTO grievances
               (name, roll_no, email, category, subject, message, created_at)
               VALUES (?,?,?,?,?,?,?)""",
            (name, clean(f.get("roll_no")), email, clean(f.get("category")),
             subject, message, datetime.now().isoformat(timespec="seconds")),
        )
        db.commit()
        flash("Grievance received. Your grievance ID is <strong>G%d</strong>. "
              "The Counselling Cell will respond within 3 working days."
              % cur.lastrowid, "success")
        return redirect(url_for("grievance"))

    return render_template("grievance.html", form={}, active="grievance")


# ------------------------------------------------------------------- admin
def admin_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if not session.get("admin"):
            flash("Please sign in to access the admin dashboard.", "error")
            return redirect(url_for("admin_login"))
        return fn(*args, **kwargs)
    return wrapper


@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        if request.form.get("password") == ADMIN_PASSWORD:
            session["admin"] = True
            flash("Signed in to the admin dashboard.", "success")
            return redirect(url_for("admin"))
        flash("Incorrect password.", "error")
        return render_template("admin_login.html"), 401
    return render_template("admin_login.html")


@app.route("/admin/logout")
def admin_logout():
    session.pop("admin", None)
    flash("Signed out.", "success")
    return redirect(url_for("home"))


@app.route("/admin")
@admin_required
def admin():
    db = get_db()
    stats = {
        "registrations": db.execute(
            "SELECT COUNT(*) c FROM registrations").fetchone()["c"],
        "academic": db.execute(
            "SELECT COUNT(*) c FROM registrations WHERE kind='academic'"
        ).fetchone()["c"],
        "admission": db.execute(
            "SELECT COUNT(*) c FROM registrations WHERE kind='admission'"
        ).fetchone()["c"],
        "appointments": db.execute(
            "SELECT COUNT(*) c FROM appointments").fetchone()["c"],
        "grievances": db.execute(
            "SELECT COUNT(*) c FROM grievances").fetchone()["c"],
        "notices": db.execute(
            "SELECT COUNT(*) c FROM notices").fetchone()["c"],
    }
    registrations = db.execute(
        "SELECT * FROM registrations ORDER BY id DESC LIMIT 300").fetchall()
    appointments = db.execute(
        "SELECT * FROM appointments ORDER BY date, slot, id DESC LIMIT 300"
    ).fetchall()
    grievances = db.execute(
        "SELECT * FROM grievances ORDER BY id DESC LIMIT 300").fetchall()
    notices = db.execute(
        "SELECT * FROM notices ORDER BY published_on DESC, id DESC"
    ).fetchall()
    return render_template("admin.html", stats=stats,
                           registrations=registrations,
                           appointments=appointments,
                           grievances=grievances, notices=notices,
                           active="admin")


@app.route("/admin/notices/add", methods=["POST"])
@admin_required
def admin_add_notice():
    f = request.form
    title = clean(f.get("title"))
    body = clean(f.get("body"))
    if not title or not body:
        flash("Notice title and body are required.", "error")
        return redirect(url_for("admin"))
    published = clean(f.get("published_on")) or date.today().isoformat()
    db = get_db()
    db.execute(
        "INSERT INTO notices (title, body, tag, published_on, created_at)"
        " VALUES (?,?,?,?,?)",
        (title, body, clean(f.get("tag")) or "General", published,
         datetime.now().isoformat(timespec="seconds")),
    )
    db.commit()
    flash("Notice published.", "success")
    return redirect(url_for("admin"))


@app.route("/admin/notices/delete/<int:notice_id>", methods=["POST"])
@admin_required
def admin_delete_notice(notice_id):
    db = get_db()
    db.execute("DELETE FROM notices WHERE id = ?", (notice_id,))
    db.commit()
    flash("Notice deleted.", "success")
    return redirect(url_for("admin"))


@app.route("/admin/appointments/status/<int:appt_id>", methods=["POST"])
@admin_required
def admin_appointment_status(appt_id):
    status = clean(request.form.get("status"))
    if status not in ("Pending", "Confirmed", "Completed", "Cancelled"):
        abort(400)
    db = get_db()
    db.execute("UPDATE appointments SET status = ? WHERE id = ?",
               (status, appt_id))
    db.commit()
    flash("Appointment #%d marked as %s." % (appt_id, status), "success")
    return redirect(url_for("admin"))


EXPORT_TABLES = ("registrations", "appointments", "grievances", "notices",
                 "oec_students", "oec_courses")


@app.route("/admin/export/<table>")
@admin_required
def admin_export(table):
    if table not in EXPORT_TABLES:
        abort(404)
    db = get_db()
    cur = db.execute("SELECT * FROM %s ORDER BY id" % table)
    cols = [d[0] for d in cur.description]
    out = io.StringIO()
    writer = csv.writer(out)
    writer.writerow(cols)
    for row in cur.fetchall():
        writer.writerow([row[c] for c in cols])
    return Response(
        out.getvalue(),
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition":
                 "attachment; filename=%s_%s.csv" % (table, date.today())},
    )


# =====================================================================
#  COUNSELLING DATA ENTRY — "Botany Counselling" interface   [Admin]
# =====================================================================
def _norm_sem(value):
    v = clean(value).upper()
    return {"5": "V", "V": "V", "SEM V": "V", "V SEM": "V", "5TH": "V",
            "SEM 5": "V", "SEMESTER V": "V",
            "6": "VI", "VI": "VI", "SEM VI": "VI", "VI SEM": "VI",
            "6TH": "VI", "SEM 6": "VI", "SEMESTER VI": "VI"}.get(v, "")


def _norm_type(value):
    v = clean(value).upper()
    return {"VAC": "VAC", "VA": "VAC", "VOC": "VOC", "VO": "VOC",
            "MJ": "MJ", "MO": "MO", "NO": "NO", "XO": "XO"}.get(v, "")


def _label_for_slot(slot_key):
    return OEC_SLOT_MAP[slot_key][3] if slot_key in OEC_SLOT_MAP else ""


def _combo(code, title):
    code, title = code or "", title or ""
    if code and title:
        return "%s — %s" % (code, title)
    return code or title


def _parse_course_cell(value):
    """'WLBLVA5001 — Fundamentals of Environment' -> 'WLBLVA5001'."""
    v = clean(value)
    if not v:
        return ""
    for sep in ("—", "–", " - "):
        if sep in v:
            v = v.split(sep)[0].strip()
            break
    return v


def _parse_counselling_date(value):
    """Accept ISO or DD-MM-YYYY; return ISO or ''."""
    v = clean(value)
    if not v:
        return ""
    if re.match(r"^\d{4}-\d{2}-\d{2}$", v):
        return v
    m = re.match(r"^(\d{1,2})[-/.](\d{1,2})[-/.](\d{4})$", v)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try:
            return date(y, mo, d).isoformat()
        except ValueError:
            return "__INVALID__"
    return "__INVALID__"


def _fmt_dmy(iso):
    iso = iso or ""
    if re.match(r"^\d{4}-\d{2}-\d{2}$", iso):
        y, m, d = iso.split("-")
        return "%s-%s-%s" % (d, m, y)
    return iso


def _oec_course_lookup(code):
    if not code:
        return None
    return get_db().execute(
        "SELECT * FROM oec_courses WHERE UPPER(course_code) = UPPER(?)",
        (code,)).fetchone()


def _oec_courses_by_slot():
    db = get_db()
    out = {}
    for key, typ, sem, _label in OEC_SLOTS:
        out[key] = db.execute(
            "SELECT * FROM oec_courses WHERE semester=? AND course_type=?"
            " ORDER BY course_code", (sem, typ)).fetchall()
    return out


def _oec_fetch_students(q="", slot="", course=""):
    sql = "SELECT * FROM oec_students"
    conds, params = [], []
    if q:
        conds.append("(LOWER(student_name) LIKE ? OR LOWER(enrolment_no)"
                     " LIKE ? OR LOWER(faculty_no) LIKE ? OR mobile LIKE ?"
                     " OR LOWER(vac_v_code) LIKE ? OR LOWER(voc_v_code)"
                     " LIKE ? OR LOWER(vac_vi_code) LIKE ? OR"
                     " LOWER(voc_vi_code) LIKE ?)")
        like = "%" + q.lower() + "%"
        params += [like] * 8
    if slot in OEC_SLOT_MAP:
        col = "%s_code" % slot
        if course:
            conds.append("%s = ?" % col)
            params.append(course)
        else:
            conds.append("%s != '' AND %s IS NOT NULL" % (col, col))
    if conds:
        sql += " WHERE " + " AND ".join(conds)
    sql += " ORDER BY id"
    return get_db().execute(sql, params).fetchall()


def _oec_filterline():
    parts = []
    if clean(request.args.get("slot")) in OEC_SLOT_MAP:
        parts.append("Sheet: " + _label_for_slot(clean(request.args.get("slot"))))
    if clean(request.args.get("course")):
        parts.append("Course: " + clean(request.args.get("course")))
    if clean(request.args.get("q")):
        parts.append("Search: " + clean(request.args.get("q")))
    return ("Filters — " + " · ".join(parts)) if parts else "All students"


def _oec_validate_student(src, db, exclude_id=None):
    """Normalise + validate one student dict. Returns (values, errors)."""
    vals = {k: clean(src.get(k)) for k in
            ("enrolment_no", "student_name", "faculty_no", "mobile",
             "semester", "counselling_date", "status", "remarks")}
    errors = []

    if not vals["student_name"]:
        errors.append("Student Name is required.")
    vals["semester"] = _norm_sem(vals["semester"])
    if src.get("semester") and not vals["semester"]:
        errors.append("Semester must be V or VI.")
    if vals["mobile"]:
        digits = "".join(ch for ch in vals["mobile"] if ch.isdigit())
        if len(digits) < 10:
            errors.append("Mobile number must have at least 10 digits.")
        else:
            vals["mobile"] = digits[-10:]
    dt = _parse_counselling_date(vals["counselling_date"])
    if dt == "__INVALID__":
        errors.append("Counselling date must be a valid date "
                      "(DD-MM-YYYY).")
        dt = clean(vals["counselling_date"]) if re.match(
            r"^\d{4}-\d{2}-\d{2}$", clean(vals["counselling_date"])) else ""
    vals["counselling_date"] = dt
    vals["status"] = vals["status"] if vals["status"] in OEC_STATUSES \
        else "Pending"
    if vals["enrolment_no"]:
        dupe = db.execute(
            "SELECT id FROM oec_students WHERE UPPER(enrolment_no) = UPPER(?)",
            (vals["enrolment_no"],)).fetchone()
        if dupe and (exclude_id is None or dupe["id"] != exclude_id):
            errors.append("Enrollment No “%s” already exists (record E%d)."
                          % (vals["enrolment_no"], dupe["id"]))
    return vals, errors


def _oec_insert_student(db, vals, slots):
    """slots: {slot_key: (code, title)}"""
    now = datetime.now().isoformat(timespec="seconds")
    cols = ["enrolment_no", "student_name", "faculty_no", "mobile",
            "semester", "counselling_date", "status", "remarks"]
    data = [vals[k] for k in cols]
    for key in OEC_SLOT_MAP:
        code, title = slots.get(key, ("", ""))
        cols += ["%s_code" % key, "%s_title" % key]
        data += [code, title]
    cols += ["created_at", "updated_at"]
    data += [now, now]
    cur = db.execute(
        "INSERT INTO oec_students (%s) VALUES (%s)"
        % (", ".join(cols), ",".join("?" * len(cols))), data)
    return cur.lastrowid


# ------------------------------------------------------------- main sheet
def _oec_main_render(form=None, status=200):
    form = dict(form or {})
    # quick-add prefill from sheet cards: /oec/?ctype=VAC VI&code=WLBLVA6002
    if request.method == "GET":
        if clean(request.args.get("ctype")):
            form.setdefault("course_type", clean(request.args.get("ctype")))
        if clean(request.args.get("code")):
            form.setdefault("course_code",
                            clean(request.args.get("code")).upper())
    return (render_template(
                "oec_main.html",
                students=_oec_fetch_students(clean(request.args.get("q"))),
                courses_by_slot=_oec_courses_by_slot(),
                form=form, active="oec", oec_tab="main"),
            status)


@app.route("/oec/")
@admin_required
def oec_main():
    return _oec_main_render()


@app.route("/oec/add", methods=["POST"])
@admin_required
def oec_add():
    f = request.form
    db = get_db()
    src = {k: clean(f.get(k)) for k in
           ("enrolment_no", "student_name", "faculty_no", "mobile",
            "semester", "counselling_date", "status", "remarks")}
    vals, errors = _oec_validate_student(src, db)

    # optional first course pick
    slots = {}
    ctype = clean(f.get("course_type"))          # e.g. "VAC V"
    slot_key = next((k for k, t, s, lbl in OEC_SLOTS
                     if lbl == ctype), "")
    code = _parse_course_cell(f.get("course_code")) or \
        _parse_course_cell(f.get("course_name_pick"))
    if code:
        c = _oec_course_lookup(code)
        if c is None:
            errors.append("Course code “%s” is not in the Course Code "
                          "Sheets — add it there first." % code)
        else:
            exp_key = next((k for k, t, s, _l in OEC_SLOTS
                            if t == c["course_type"] and s == c["semester"]),
                           slot_key)
            slots[exp_key] = (c["course_code"], c["title"])

    if errors:
        for e in errors:
            flash(e, "error")
        return _oec_main_render(form=src, status=400)

    sid = _oec_insert_student(db, vals, slots)
    db.commit()
    flash("Student <strong>E%d — %s</strong> saved to the Main Entry Sheet."
          " Course sheets are updated automatically." % (sid, vals["student_name"]),
          "success")
    return redirect(url_for("oec_main"))


# ------------------------------------------------------ live JSON updates
def _json_error(msg, status=400):
    return jsonify(ok=False, error=msg), status


@app.route("/oec/api/student/<int:sid>/field", methods=["POST"])
@admin_required
def oec_api_field(sid):
    db = get_db()
    row = db.execute("SELECT * FROM oec_students WHERE id=?",
                     (sid,)).fetchone()
    if row is None:
        return _json_error("Record not found.", 404)
    field = clean(request.form.get("field"))
    if field not in ("enrolment_no", "student_name", "faculty_no",
                     "mobile", "semester", "counselling_date",
                     "status", "remarks"):
        return _json_error("Unknown field.")
    value = clean(request.form.get("value"))
    if field == "student_name" and not value:
        return _json_error("Student Name cannot be empty.")
    if field == "semester":
        value = _norm_sem(value)
        if not value:
            return _json_error("Semester must be V or VI.")
    if field == "status" and value not in OEC_STATUSES:
        return _json_error("Unknown status.")
    if field == "counselling_date":
        value = _parse_counselling_date(value)
        if value == "__INVALID__":
            return _json_error("Invalid date — use DD-MM-YYYY.")
    if field == "mobile" and value:
        digits = "".join(ch for ch in value if ch.isdigit())
        if len(digits) < 10:
            return _json_error("Mobile must have at least 10 digits.")
        value = digits[-10:]
    if field == "enrolment_no" and value:
        dupe = db.execute(
            "SELECT id FROM oec_students WHERE UPPER(enrolment_no)=UPPER(?)"
            " AND id != ?", (value, sid)).fetchone()
        if dupe:
            return _json_error("Enrollment No already exists (E%d)."
                               % dupe["id"])
    db.execute(
        "UPDATE oec_students SET %s=?, updated_at=? WHERE id=?" % field,
        (value, datetime.now().isoformat(timespec="seconds"), sid))
    db.commit()
    return jsonify(ok=True)


@app.route("/oec/api/student/<int:sid>/slot", methods=["POST"])
@admin_required
def oec_api_slot(sid):
    db = get_db()
    row = db.execute("SELECT * FROM oec_students WHERE id=?",
                     (sid,)).fetchone()
    if row is None:
        return _json_error("Record not found.", 404)
    slot = clean(request.form.get("slot"))
    if slot not in OEC_SLOT_MAP:
        return _json_error("Unknown slot.")
    code = _parse_course_cell(request.form.get("code"))
    title = ""
    if code:
        c = _oec_course_lookup(code)
        if c is None:
            return _json_error("Unknown course code “%s”. Add it in Course "
                               "Code Sheets first." % code)
        exp_type, exp_sem = OEC_SLOT_MAP[slot][1], OEC_SLOT_MAP[slot][2]
        if c["course_type"] != exp_type or c["semester"] != exp_sem:
            return _json_error("%s is a %s (Sem %s) course — wrong column."
                               % (code, c["course_type"], c["semester"]))
        title = c["title"]
        code = c["course_code"]
    db.execute(
        "UPDATE oec_students SET %s_code=?, %s_title=?, updated_at=?"
        " WHERE id=?" % (slot, slot),
        (code, title, datetime.now().isoformat(timespec="seconds"), sid))
    db.commit()
    return jsonify(ok=True, value=_combo(code, title))


@app.route("/oec/api/student/<int:sid>/delete", methods=["POST"])
@admin_required
def oec_api_delete(sid):
    db = get_db()
    db.execute("DELETE FROM oec_students WHERE id=?", (sid,))
    db.commit()
    return jsonify(ok=True)


@app.route("/oec/api/bulk-delete", methods=["POST"])
@admin_required
def oec_api_bulk_delete():
    ids = [int(x) for x in re.findall(r"\d+",
                                      request.form.get("ids", ""))]
    if not ids:
        return _json_error("No records selected.")
    db = get_db()
    db.execute("DELETE FROM oec_students WHERE id IN (%s)"
               % ",".join("?" * len(ids)), ids)
    db.commit()
    return jsonify(ok=True, deleted=len(ids))


# -------------------------------------------------------- per-slot sheets
@app.route("/oec/sheet/<slot>")
@admin_required
def oec_sheet(slot):
    """Per-paper-code sheets: every code of the slot gets its own sheet,
    even with zero students ('No students in CODE yet.')."""
    if slot not in OEC_SLOT_MAP:
        abort(404)
    typ, sem = OEC_SLOT_MAP[slot][1], OEC_SLOT_MAP[slot][2]
    db = get_db()
    all_codes = db.execute(
        "SELECT DISTINCT course_code FROM oec_courses WHERE semester=?"
        " AND course_type=? ORDER BY course_code", (sem, typ)).fetchall()
    groups = {c["course_code"]: [] for c in all_codes}
    for s in _oec_fetch_students(slot=slot):
        groups.setdefault(s["%s_code" % slot], []).append(s)
    course = clean(request.args.get("course")).upper()
    counts = {code: len(rows) for code, rows in groups.items()}
    ordered = sorted(groups.items())
    if course:
        ordered = [(c, r) for c, r in ordered if c == course]
    return render_template("oec_sheet.html", slot=slot,
                           label=_label_for_slot(slot),
                           dash_label="%s-%s" % (typ, sem.replace("VI", "VI")),
                           groups=ordered, counts=counts,
                           total_codes=len(groups), course=course,
                           active="oec", oec_tab=slot)


@app.route("/oec/student/<int:sid>.qr.png")
@admin_required
def oec_student_qr(sid):
    """QR identity card code for a counselling record."""
    if not HAS_QR:
        abort(404)
    s = get_db().execute("SELECT * FROM oec_students WHERE id=?",
                         (sid,)).fetchone()
    if s is None:
        abort(404)
    slots = ";".join(_combo(s["%s_code" % k], "").split(" — ")[0]
                     for k in OEC_SLOT_MAP if s["%s_code" % k])
    payload = ("AMU-BSC-COUNSELLING|E%d|%s|%s|%s|%s|%s"
               % (sid, s["enrolment_no"], s["student_name"],
                  s["semester"] or "", slots,
                  _fmt_dmy(s["counselling_date"])))
    import qrcode as _qr
    img = _qr.make(payload, error_correction=_qr.constants.ERROR_CORRECT_M)
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    return send_file(buf, mimetype="image/png")


def _build_course_xlsx(label, code, title, rows):
    """Single-sheet workbook for one paper code."""
    from openpyxl import Workbook
    from openpyxl.styles import Font as XFont
    wb = Workbook()
    ws = wb.active
    ws.title = ("%s %s" % (label.replace(" ", "-"), code))[:31]
    subtitle = ("Aligarh Muslim University, Aligarh · B.Sc. Student Portal ·"
                " Session 2026-27 · Generated " + date.today().isoformat())
    _xls_title_rows(ws, ("%s — %s" % (code, title) if title else code) +
                    " (%s · %d student(s))" % (label, len(rows)),
                    subtitle, len(SPLIT_HEADERS))
    for ci, h in enumerate(SPLIT_HEADERS, 1):
        ws.cell(row=4, column=ci, value=h)
    _xls_header_row(ws, 4, len(SPLIT_HEADERS))
    _xls_set_widths(ws, SPLIT_WIDTHS)
    ws.freeze_panes = "A5"
    values = [[i, s["enrolment_no"], s["student_name"], s["faculty_no"],
               s["mobile"], code, title, _fmt_dmy(s["counselling_date"]),
               s["status"], s["remarks"]]
              for i, s in enumerate(rows, 1)]
    end = _xls_write(ws, values, 5)
    tc = ws.cell(row=end + 1, column=3,
                 value="Total: %d student(s)" % len(rows))
    tc.font = XFont(bold=True)
    return wb


@app.route("/oec/sheet/<slot>/<code>.xlsx")
@admin_required
def oec_course_xlsx(slot, code):
    if slot not in OEC_SLOT_MAP:
        abort(404)
    code = code.upper()
    rows = _oec_fetch_students(slot=slot, course=code)
    title = rows[0]["%s_title" % slot] if rows else ""
    return _xls_send(_build_course_xlsx(_label_for_slot(slot), code,
                                        title, rows),
                     "%s_%s_%s.xlsx"
                     % (_label_for_slot(slot).replace(" ", "_"), code,
                        date.today()))


# ------------------------------------------------------------ import file
def _stu_key_for(header):
    h = re.sub(r"[^a-z0-9]+", " ", (header or "").lower()).strip()
    # literal column names first (NOTE: "S.No" contains "no" — exclude it
    # before slot-token matching)
    plain = {"s no": None, "sno": None, "sl no": None, "serial": None,
             "no": None, "sl": None, "#": None,
             "enrolment no": "enrolment_no", "enrollment no": "enrolment_no",
             "enrolment": "enrolment_no", "enrollment": "enrolment_no",
             "student name": "student_name", "name": "student_name",
             "faculty no": "faculty_no", "faculty number": "faculty_no",
             "fac no": "faculty_no",
             "mobile": "mobile", "mobile no": "mobile", "phone": "mobile",
             "semester": "semester", "sem": "semester",
             "counselling date": "counselling_date",
             "counseling date": "counselling_date",
             "date": "counselling_date",
             "status": "status", "remarks": "remarks", "remark": "remarks"}
    if h in plain:
        return plain[h]
    m = re.search(r"\b(vac|voc|mj|mo|no|xo)\b", h)
    if m:
        sem = "vi" if re.search(r"\bvi\b", h) else "v"
        key = "%s_%s" % (m.group(1), sem)
        return key if key in OEC_SLOT_MAP else None
    return None


@app.route("/oec/import", methods=["GET", "POST"])
@admin_required
def oec_import():
    if request.method == "POST":
        file = request.files.get("file")
        if file is None or not file.filename:
            flash("Please choose an .xlsx or .csv file to import.", "error")
            return render_template("oec_import.html", active="oec",
                                   oec_tab=""), 400
        name = file.filename.lower()
        if name.endswith(".csv"):
            raw = file.read().decode("utf-8-sig", errors="replace")
            rows = [[clean(c) for c in row]
                    for row in csv.reader(io.StringIO(raw))]
        elif name.endswith((".xlsx", ".xlsm")):
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(file.read()), read_only=True,
                               data_only=True)
            ws = wb["Data Entry"] if "Data Entry" in wb.sheetnames \
                else wb.active
            rows = [["" if c is None else clean(str(c)) for c in r]
                    for r in ws.iter_rows(values_only=True)]
        else:
            flash("Unsupported file type — upload .xlsx or .csv.", "error")
            return render_template("oec_import.html", active="oec",
                                   oec_tab=""), 400

        header_idx, colmap = None, {}
        for idx, row in enumerate(rows[:10]):
            hits, cmap = 0, {}
            for ci, cell in enumerate(row):
                key = _stu_key_for(cell)
                if key:
                    cmap[ci] = key
                    hits += 1
            if hits >= 3:
                header_idx, colmap = idx, cmap
                break
        if header_idx is None:
            flash("Header row not detected. Use the official Excel format "
                  "(download from Main Entry Sheet).", "error")
            return render_template("oec_import.html", active="oec",
                                   oec_tab=""), 400

        db = get_db()
        ok, errors = 0, []
        for i, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
            if not any(row):
                continue
            src, slots = {}, {}
            for ci, key in colmap.items():
                val = row[ci] if ci < len(row) else ""
                if key in OEC_SLOT_MAP:
                    code = _parse_course_cell(val)
                    if code:
                        slots[key] = code
                else:
                    src[key] = val
            vals, verrs = _oec_validate_student(src, db)
            slot_map = {}
            for key, code in slots.items():
                c = _oec_course_lookup(code)
                if c is None:
                    verrs.append("Unknown course code “%s” in %s"
                                 % (code, _label_for_slot(key)))
                else:
                    exp_type, exp_sem = OEC_SLOT_MAP[key][1], \
                        OEC_SLOT_MAP[key][2]
                    if c["course_type"] != exp_type \
                            or c["semester"] != exp_sem:
                        verrs.append("“%s” belongs to %s Sem %s, not %s"
                                     % (code, c["course_type"],
                                        c["semester"], _label_for_slot(key)))
                    else:
                        slot_map[key] = (c["course_code"], c["title"])
            if verrs:
                errors.append("Row %d: %s" % (i, "; ".join(verrs)))
            else:
                _oec_insert_student(db, vals, slot_map)
                ok += 1
        db.commit()
        total = ok + len(errors)
        if ok:
            flash("Imported <strong>%d of %d</strong> student(s) from “%s”."
                  % (ok, total, file.filename), "success")
        else:
            flash("No rows imported from “%s”." % file.filename, "error")
        for e in errors[:5]:
            flash(e, "error")
        if len(errors) > 5:
            flash("…and %d more rejected row(s)." % (len(errors) - 5),
                  "error")
        return redirect(url_for("oec_import"))

    return render_template("oec_import.html", active="oec", oec_tab="")


# ----------------------------------------------------- excel workbooks
def _xls_header_row(ws, row_idx, ncols):
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    fill = PatternFill("solid", fgColor="0B5D3B")
    font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="B9CFC2")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ci in range(1, ncols + 1):
        cell = ws.cell(row=row_idx, column=ci)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = border


def _xls_title_rows(ws, title, subtitle, ncols):
    from openpyxl.styles import Font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    t = ws.cell(row=1, column=1, value=title)
    t.font = Font(bold=True, size=14, color="0B5D3B")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    s = ws.cell(row=2, column=1, value=subtitle)
    s.font = Font(italic=True, size=10, color="5D7368")


def _xls_set_widths(ws, widths):
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _xls_write(ws, rows_values, start_row):
    from openpyxl.styles import Alignment, Border, Side
    thin = Side(style="thin", color="D9E5DD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for values in rows_values:
        for ci, v in enumerate(values, start=1):
            cell = ws.cell(row=start_row, column=ci, value=v)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        start_row += 1
    return start_row


def _stu_row_values(sno, s):
    return ([sno, s["enrolment_no"], s["student_name"], s["faculty_no"],
             s["mobile"], s["semester"]]
            + [_combo(s["%s_code" % key], s["%s_title" % key])
               for key in OEC_SLOT_MAP]
            + [_fmt_dmy(s["counselling_date"]), s["status"], s["remarks"]])


def _xls_send(wb, filename):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf, as_attachment=True, download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument"
                 ".spreadsheetml.sheet")


def _build_main_xlsx(students, subtitle):
    from openpyxl import Workbook
    from openpyxl.styles import Font
    wb = Workbook()
    ws = wb.active
    ws.title = "Main Entry Sheet"
    ncols = len(STU_SHEET_HEADERS)
    _xls_title_rows(ws, "MAIN ENTRY SHEET — OPEN ELECTIVE (VAC/VOC)",
                    subtitle, ncols)
    for ci, h in enumerate(STU_SHEET_HEADERS, 1):
        ws.cell(row=4, column=ci, value=h)
    _xls_header_row(ws, 4, ncols)
    _xls_set_widths(ws, STU_WIDTHS)
    ws.freeze_panes = "A5"
    end = _xls_write(ws, [_stu_row_values(i + 1, s)
                          for i, s in enumerate(students)], 5)
    ws.auto_filter.ref = "A4:M%d" % max(4, end - 1)
    from openpyxl.styles import Font as F
    tc = ws.cell(row=end + 1, column=3,
                 value="Total: %d student(s)" % len(students))
    tc.font = F(bold=True)
    return wb


def _build_split_xlsx(students):
    """One workbook: Summary + sheet per (slot, course) + Main (All)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font as XFont
    from openpyxl.utils import get_column_letter

    subtitle = ("Aligarh Muslim University, Aligarh · B.Sc. Student "
                "Portal · Session 2026-27 · Generated " + date.today().isoformat())
    groups = {}
    for s in students:
        for key, _t, _s_, label in OEC_SLOTS:
            code = s["%s_code" % key]
            if code:
                groups.setdefault((key, code), []).append(s)

    def order(item):
        (key, code), _rows = item
        return ([k for k, _t, _s, _l in OEC_SLOTS].index(key), code)

    ordered = sorted(groups.items(), key=order)
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    heads = ["Sheet", "Course Code", "Course Title", "Students"]
    _xls_title_rows(ws, "COURSE-WISE SPLIT — SUMMARY", subtitle, len(heads))
    for ci, h in enumerate(heads, 1):
        ws.cell(row=4, column=ci, value=h)
    _xls_header_row(ws, 4, len(heads))
    for i, wd in enumerate([14, 16, 52, 10], 1):
        ws.column_dimensions[get_column_letter(i)].width = wd
    r = 5
    for (key, code), rows_ in ordered:
        ws.cell(row=r, column=1, value=_label_for_slot(key))
        ws.cell(row=r, column=2, value=code)
        ws.cell(row=r, column=3, value=rows_[0]["%s_title" % key] or "")
        ws.cell(row=r, column=4, value=len(rows_))
        r += 1
    if ordered:
        c1 = ws.cell(row=r, column=3, value="TOTAL")
        c1.font = XFont(bold=True)
        c2 = ws.cell(row=r, column=4, value=sum(len(x[1]) for x in ordered))
        c2.font = XFont(bold=True)
    else:
        ws.cell(row=r, column=1, value="No course allotments yet.")
    ws.freeze_panes = "A5"

    for (key, code), rows_ in ordered:
        name = ("%s %s" % (_label_for_slot(key).replace(" ", "-"), code))[:31]
        sh = wb.create_sheet(name)
        title = "%s — %s (%s · %d student(s))" % (
            code, rows_[0]["%s_title" % key] or "",
            _label_for_slot(key), len(rows_))
        _xls_title_rows(sh, title, subtitle, len(SPLIT_HEADERS))
        for ci, h in enumerate(SPLIT_HEADERS, 1):
            sh.cell(row=4, column=ci, value=h)
        _xls_header_row(sh, 4, len(SPLIT_HEADERS))
        _xls_set_widths(sh, SPLIT_WIDTHS)
        sh.freeze_panes = "A5"
        values = []
        for i, s in enumerate(rows_, 1):
            values.append([i, s["enrolment_no"], s["student_name"],
                           s["faculty_no"], s["mobile"], code,
                           s["%s_title" % key] or "",
                           _fmt_dmy(s["counselling_date"]), s["status"],
                           s["remarks"]])
        _xls_write(sh, values, 5)

    allws = wb.create_sheet("Main Sheet (All)")
    _xls_title_rows(allws, "MAIN ENTRY SHEET (ALL DATA)", subtitle,
                    len(STU_SHEET_HEADERS))
    for ci, h in enumerate(STU_SHEET_HEADERS, 1):
        allws.cell(row=4, column=ci, value=h)
    _xls_header_row(allws, 4, len(STU_SHEET_HEADERS))
    _xls_set_widths(allws, STU_WIDTHS)
    allws.freeze_panes = "A5"
    _xls_write(allws, [_stu_row_values(i + 1, s)
                       for i, s in enumerate(students)], 5)
    return wb


def _build_format_xlsx(courses, with_samples=False):
    from openpyxl import Workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.styles import Font
    wb = Workbook()
    ws = wb.active
    ws.title = "Data Entry"
    for ci, h in enumerate(STU_SHEET_HEADERS, 1):
        ws.cell(row=1, column=ci, value=h)
    _xls_header_row(ws, 1, len(STU_SHEET_HEADERS))
    _xls_set_widths(ws, STU_WIDTHS)
    ws.freeze_panes = "A2"

    # dropdown validations — F=Semester, G..=course slots, then
    # Counselling Date, Status, Remarks (positions follow len(OEC_SLOTS))
    from openpyxl.utils import get_column_letter
    dv_sem = DataValidation(type="list", formula1='"V,VI"',
                            allow_blank=True)
    dv_status = DataValidation(
        type="list",
        formula1='"%s"' % ",".join(OEC_STATUSES), allow_blank=True)
    dv_code = DataValidation(type="list",
                             formula1="'Course Code Sheets'!$C$2:$C$500",
                             allow_blank=True)
    ws.add_data_validation(dv_sem)
    ws.add_data_validation(dv_status)
    ws.add_data_validation(dv_code)
    dv_sem.add("F2:F2000")
    n = len(OEC_SLOTS)
    for i in range(n):
        col = get_column_letter(7 + i)                       # G onwards
        dv_code.add("%s2:%s2000" % (col, col))
    status_col = get_column_letter(8 + n)
    dv_status.add("%s2:%s2000" % (status_col, status_col))

    if with_samples:
        ws.append([1, "2026BOT001", "AMAN KUMAR", "BSC2026/45",
                   "9876543210", "V", "WLBLVA5001", "ZYBLVO5007",
                   "BTBLMO5003", "", "", "", "",
                   "05-07-2026", "Counselling Done", "Sample row"])
        ws.append([2, "2026BOT002", "SANA PARVEEN", "BSC2026/46",
                   "9876501234", "V", "", "", "", "BCBLNO5005",
                   "CHBSXO5003", "", "",
                   "07-07-2026", "Pending", ""])
        ws.append([3, "2026BOT003", "ROHAN VERMA", "BSC2026/47",
                   "9812345678", "VI", "", "", "", "", "",
                   "WLBLVA6002", "",
                   "10-07-2026", "Pending", "Paid"])

    cm = wb.create_sheet("Course Code Sheets")
    for ci, h in enumerate(["Semester", "Course Type", "Course Code",
                            "Course Title"], 1):
        cm.cell(row=1, column=ci, value=h)
    _xls_header_row(cm, 1, 4)
    from openpyxl.utils import get_column_letter
    for i, wd in enumerate([10, 12, 16, 52], 1):
        cm.column_dimensions[get_column_letter(i)].width = wd
    for ri, c in enumerate(courses, start=2):
        cm.cell(row=ri, column=1, value=c["semester"])
        cm.cell(row=ri, column=2, value=c["course_type"])
        cm.cell(row=ri, column=3, value=c["course_code"])
        cm.cell(row=ri, column=4, value=c["title"])

    ins = wb.create_sheet("Instructions")
    lines = [
        "MAIN ENTRY SHEET FORMAT · B.Sc. Sem V & VI · Open Elective (VAC/VOC) · Session 2026-27",
        "",
        "1. Enter data ONLY in the 'Data Entry' sheet from row 2 (row 1 = headers — do not change).",
        "2. 'Student Name' is compulsory. Enrollment No should be unique.",
        "3. Course columns: type the CODE alone or 'CODE — Course Name' (e.g. WLBLVA5001 — Fundamentals of Environment).",
        "   A VAC-V course goes in 'VAC V (Course Code)' — wrong-column entries are rejected at import.",
        "4. Counselling Date format: DD-MM-YYYY (e.g. 05-07-2026).",
        "5. Status: " + " / ".join(OEC_STATUSES) + ".",
        "6. Save as .xlsx (or .csv) and import from Main Entry Sheet → Import Excel / Import CSV.",
        "",
        "After import, the portal automatically separates records course-wise —",
        "download them from Main Entry Sheet → Excel / ZIP / Split Sheets.",
    ]
    for ri, line in enumerate(lines, 1):
        cell = ins.cell(row=ri, column=1, value=line)
        if ri == 1:
            cell.font = Font(bold=True, size=12, color="0B5D3B")
    ins.column_dimensions["A"].width = 110
    return wb


# ------------------------------------------------------------- downloads
@app.route("/oec/format.xlsx")
@admin_required
def oec_format():
    courses = get_db().execute(
        "SELECT * FROM oec_courses ORDER BY semester, course_type,"
        " course_code").fetchall()
    return _xls_send(_build_format_xlsx(courses),
                     "Main_Entry_Format_%s.xlsx" % date.today())


@app.route("/oec/sample.xlsx")
@admin_required
def oec_sample():
    courses = get_db().execute(
        "SELECT * FROM oec_courses ORDER BY semester, course_type,"
        " course_code").fetchall()
    return _xls_send(_build_format_xlsx(courses, with_samples=True),
                     "Main_Entry_Sample_%s.xlsx" % date.today())


@app.route("/oec/export.xlsx")
@admin_required
def oec_export():
    q = clean(request.args.get("q"))
    slot = clean(request.args.get("slot"))
    course = clean(request.args.get("course"))
    students = _oec_fetch_students(q=q, slot=slot, course=course)
    subtitle = "%s · Session 2026-27 · %s · Generated %s" % (
        "Aligarh Muslim University, Aligarh", _oec_filterline(),
        date.today().isoformat())
    return _xls_send(_build_main_xlsx(students, subtitle),
                     "Main_Entry_Sheet_%s.xlsx" % date.today())


@app.route("/oec/split.xlsx")
@admin_required
def oec_split():
    students = _oec_fetch_students(
        q=clean(request.args.get("q")),
        slot=clean(request.args.get("slot")),
        course=clean(request.args.get("course")))
    return _xls_send(_build_split_xlsx(students),
                     "Coursewise_Sheets_%s.xlsx" % date.today())


@app.route("/oec/export.zip")
@admin_required
def oec_export_zip():
    db = get_db()
    students = _oec_fetch_students()
    buf = io.BytesIO()
    subtitle = ("Aligarh Muslim University, Aligarh · Session 2026-27 · "
                "Generated " + date.today().isoformat())
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        b = io.BytesIO()
        _build_main_xlsx(students, subtitle).save(b)
        z.writestr("Main_Entry_Sheet.xlsx", b.getvalue())
        b = io.BytesIO()
        _build_split_xlsx(students).save(b)
        z.writestr("Coursewise_Split_Sheets.xlsx", b.getvalue())
        for key, _t, _s, label in OEC_SLOTS:
            rows_ = [s for s in students if s["%s_code" % key]]
            if not rows_:
                continue
            b = io.BytesIO()
            _build_split_xlsx(rows_).save(b)
            z.writestr("%s_Sheet.xlsx" % label.replace(" ", "_"),
                       b.getvalue())
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name="Counselling_Sheets_%s.zip"
                     % date.today(), mimetype="application/zip")


# --------------------------------------------------------------- PDF out
def _pdf_main(students, subtitle, focus=None):
    """Printable PDF. With focus=slot_key -> compact per-course layout."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (SimpleDocTemplate, Spacer, Table,
                                    TableStyle, Paragraph)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=9 * mm, rightMargin=9 * mm,
                            topMargin=9 * mm, bottomMargin=9 * mm,
                            title="Main Entry Sheet")
    green = colors.HexColor("#0b5d3b")
    pcell = ParagraphStyle("cell", fontName="Helvetica", fontSize=6.7,
                           leading=8.4)
    phead = ParagraphStyle("head", fontName="Helvetica-Bold",
                           fontSize=7, leading=8.6,
                           textColor=colors.white)
    ptitle = ParagraphStyle("title", fontName="Helvetica-Bold",
                            fontSize=13, textColor=green)
    psub = ParagraphStyle("sub", fontName="Helvetica-Oblique", fontSize=8,
                          textColor=colors.HexColor("#5d7368"))
    if focus:
        # per-course compact table
        heads = ["S.No", "Enrollment No", "Student Name", "Faculty No.",
                 "Mobile", "Course (Code — Title)", "Counselling Date",
                 "Status"]
        data = [[Paragraph(h, phead) for h in heads]]
        for i, s in enumerate(students, 1):
            row = [i, s["enrolment_no"], s["student_name"], s["faculty_no"],
                   s["mobile"],
                   _combo(s["%s_code" % focus], s["%s_title" % focus]),
                   _fmt_dmy(s["counselling_date"]), s["status"]]
            data.append([Paragraph(str(v or ""), pcell) for v in row])
        widths = [10 * mm, 24 * mm, 34 * mm, 22 * mm, 22 * mm, 65 * mm,
                  22 * mm, 20 * mm]
    else:
        heads = (["S.No", "Enrollment No", "Student Name", "Faculty No.",
                  "Mobile", "Sem"]
                 + ["%s (Course Code)" % s[3] for s in OEC_SLOTS]
                 + ["Counselling Date", "Status"])
        data = [[Paragraph(h, phead) for h in heads]]
        for i, s in enumerate(students, 1):
            vals = _stu_row_values(i, s)
            vals.pop()                              # drop Remarks (width)
            data.append([Paragraph(str(v or ""), pcell) for v in vals])
        n = len(OEC_SLOTS)
        widths = ([8 * mm, 18 * mm, 24 * mm, 15 * mm, 15 * mm, 7 * mm]
                  + [13 * mm] * n + [15 * mm, 13 * mm])
    table = Table(data, colWidths=widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), green),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#b9cfc2")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.white, colors.HexColor("#f3f8f5")]),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))
    doc.build([
        Paragraph("MAIN ENTRY SHEET — OPEN ELECTIVE (VAC/VOC)", ptitle),
        Paragraph(subtitle, psub), Spacer(1, 4), table])
    buf.seek(0)
    return buf


def _pdf_slip(s):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (SimpleDocTemplate, Spacer, Table,
                                    TableStyle, Paragraph)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=18 * mm,
                            rightMargin=18 * mm, topMargin=16 * mm,
                            bottomMargin=16 * mm,
                            title="Counselling Slip E%d" % s["id"])
    green = colors.HexColor("#0b5d3b")
    ptitle = ParagraphStyle("t", fontName="Helvetica-Bold", fontSize=14,
                            textColor=green, alignment=1)
    psub = ParagraphStyle("s", fontName="Helvetica", fontSize=9,
                          textColor=colors.HexColor("#5d7368"), alignment=1)
    rows = [["Record ID", "E%d" % s["id"]],
            ["Enrollment No", s["enrolment_no"] or "—"],
            ["Student Name", s["student_name"]],
            ["Faculty No.", s["faculty_no"] or "—"],
            ["Mobile", s["mobile"] or "—"],
            ["Semester", s["semester"] or "—"],
            ["VAC V Course", _combo(s["vac_v_code"], s["vac_v_title"])
             or "—"],
            ["VOC V Course", _combo(s["voc_v_code"], s["voc_v_title"])
             or "—"],
            ["VAC VI Course", _combo(s["vac_vi_code"], s["vac_vi_title"])
             or "—"],
            ["VOC VI Course", _combo(s["voc_vi_code"], s["voc_vi_title"])
             or "—"],
            ["Counselling Date", _fmt_dmy(s["counselling_date"]) or "—"],
            ["Status", s["status"] or "Pending"],
            ["Remarks", s["remarks"] or "—"]]
    table = Table(rows, colWidths=[45 * mm, 105 * mm])
    table.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#b9cfc2")),
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#eef6f0")),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9.5),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    sig = Table([["_________________", "_________________"],
                 ["Checked by", "Convener, Counselling Cell"]],
                colWidths=[75 * mm, 75 * mm])
    sig.setStyle(TableStyle([("ALIGN", (0, 0), (-1, -1), "CENTER"),
                             ("FONTSIZE", (0, 0), (-1, -1), 9),
                             ("TEXTCOLOR", (0, 1), (-1, 1), green)]))
    doc.build([
        Paragraph("ALIGARH MUSLIM UNIVERSITY, ALIGARH", ptitle),
        Paragraph("B.Sc. Student Portal · Counselling Slip · "
                  "Session 2026-27", psub),
        Spacer(1, 10), table, Spacer(1, 34), sig])
    buf.seek(0)
    return buf


@app.route("/oec/export.pdf")
@admin_required
def oec_export_pdf():
    if not HAS_PDF:
        return redirect(url_for("oec_print", **request.args))
    slot = clean(request.args.get("slot"))
    students = _oec_fetch_students(
        q=clean(request.args.get("q")),
        slot=slot,
        course=clean(request.args.get("course")).upper())
    focus = slot if slot in OEC_SLOT_MAP else None
    buf = _pdf_main(students, "%s · %s · Generated %s" % (
        "Aligarh Muslim University, Aligarh", _oec_filterline(),
        date.today().isoformat()), focus=focus)
    return send_file(buf, as_attachment=True,
                     download_name="Main_Entry_Sheet_%s.pdf"
                     % date.today(), mimetype="application/pdf")


@app.route("/oec/student/<int:sid>.pdf")
@admin_required
def oec_student_pdf(sid):
    if not HAS_PDF:
        abort(404)
    s = get_db().execute("SELECT * FROM oec_students WHERE id=?",
                         (sid,)).fetchone()
    if s is None:
        abort(404)
    return send_file(_pdf_slip(s), as_attachment=True,
                     download_name="Counselling_Slip_E%d.pdf" % sid,
                     mimetype="application/pdf")


# ----------------------------------------------------------------- print
@app.route("/oec/print")
@admin_required
def oec_print():
    slot = clean(request.args.get("slot"))
    course = clean(request.args.get("course")).upper()
    students = _oec_fetch_students(q=clean(request.args.get("q")),
                                   slot=slot, course=course)
    focus = slot if slot in OEC_SLOT_MAP else ""
    return render_template("oec_print.html", students=students,
                           focus=focus, focus_label=_label_for_slot(slot),
                           filterline=_oec_filterline(), active="oec")


@app.route("/oec/student/<int:sid>/print")
@admin_required
def oec_student_print(sid):
    s = get_db().execute("SELECT * FROM oec_students WHERE id=?",
                         (sid,)).fetchone()
    if s is None:
        abort(404)
    return render_template("oec_slip.html", s=s, combo=_combo,
                           fmt=_fmt_dmy, active="oec")


# ------------------------------------------------------ backup / restore
@app.route("/oec/backup.db")
@admin_required
def oec_backup():
    return send_file(DB_PATH, as_attachment=True,
                     download_name="counselling_backup_%s.db"
                     % datetime.now().strftime("%Y%m%d_%H%M%S"),
                     mimetype="application/octet-stream")


@app.route("/oec/restore", methods=["POST"])
@admin_required
def oec_restore():
    file = request.files.get("file")
    if file is None or not file.filename:
        flash("Choose a .db backup file to restore.", "error")
        return redirect(url_for("oec_import"))
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".db")
    file.save(tmp.name)
    tmp.close()
    ok, msg = False, ""
    try:
        with open(tmp.name, "rb") as fh:
            if not fh.read(16).startswith(b"SQLite format 3"):
                raise ValueError("Not a SQLite database file.")
        test = sqlite3.connect(tmp.name)
        tables = {r[0] for r in test.execute(
            "SELECT name FROM sqlite_master WHERE type='table'")}
        test.close()
        if "oec_students" not in tables or "oec_courses" not in tables:
            raise ValueError("Backup is missing the counselling tables.")
        ok = True
    except Exception as exc:
        msg = str(exc)
    if not ok:
        os.unlink(tmp.name)
        flash("Restore failed: %s" % msg, "error")
        return redirect(url_for("oec_import"))
    bak = DB_PATH + ".bak-" + datetime.now().strftime("%Y%m%d%H%M%S")
    shutil.copy2(DB_PATH, bak)
    shutil.move(tmp.name, DB_PATH)
    flash("Database restored from backup. A safety copy of the previous "
          "data was saved to <code>%s</code>." % os.path.basename(bak),
          "success")
    return redirect(url_for("oec_main"))


# --------------------------------------------------------- course update
@app.route("/oec/courses")
@admin_required
def oec_courses():
    db = get_db()
    rows = db.execute(
        "SELECT * FROM oec_courses ORDER BY semester, course_type,"
        " course_code").fetchall()
    groups = []
    for sem in OEC_SEMS:
        for typ in OEC_TYPES_ALL:
            g = [c for c in rows
                 if c["semester"] == sem and c["course_type"] == typ]
            if g:
                groups.append((sem, typ, g))
    return render_template("oec_courses.html", groups=groups,
                           active="oec", oec_tab="courses")


def _oec_course_payload(f):
    sem = _norm_sem(f.get("semester"))
    typ = _norm_type(f.get("course_type"))
    code = clean(f.get("course_code")).upper()
    title = clean(f.get("title"))            # title is OPTIONAL — code is key
    errors = []
    if not sem:
        errors.append("Semester must be V or VI.")
    if not typ:
        errors.append("Course type must be VAC, VOC, MO, NO, XO or MJ.")
    if not code:
        errors.append("Course code is required.")
    return (sem, typ, code, title), errors


@app.route("/oec/courses/add", methods=["POST"])
@admin_required
def oec_course_add():
    (sem, typ, code, title), errors = _oec_course_payload(request.form)
    db = get_db()
    if not errors and db.execute(
            "SELECT 1 FROM oec_courses WHERE UPPER(course_code)=?",
            (code,)).fetchone():
        errors.append("Course code %s already exists." % code)
    if errors:
        for e in errors:
            flash(e, "error")
        return redirect(url_for("oec_courses"))
    db.execute(
        "INSERT INTO oec_courses (semester, course_type, course_code,"
        " title, created_at) VALUES (?,?,?,?,?)",
        (sem, typ, code, title, datetime.now().isoformat(timespec="seconds")))
    db.commit()
    flash("Course added: %s · %s · %s — %s" % (sem, typ, code, title),
          "success")
    return redirect(url_for("oec_courses"))


def _slot_for_course(sem, typ):
    """Which assignment slot (column pair) does a course feed, if any."""
    for key, t, s, _l in OEC_SLOTS:
        if t == typ and s == sem:
            return key
    return None


@app.route("/oec/courses/update/<int:course_id>", methods=["POST"])
@admin_required
def oec_course_update(course_id):
    db = get_db()
    old = db.execute("SELECT * FROM oec_courses WHERE id = ?",
                     (course_id,)).fetchone()
    if old is None:
        abort(404)
    (sem, typ, code, title), errors = _oec_course_payload(request.form)
    if not errors:
        dup = db.execute(
            "SELECT id FROM oec_courses WHERE UPPER(course_code)=?"
            " AND id != ?", (code, course_id)).fetchone()
        if dup:
            errors.append("Course code %s is already used by course #%d."
                          % (code, dup["id"]))
    if errors:
        for e in errors:
            flash(e, "error")
        return redirect(url_for("oec_courses"))

    db.execute(
        "UPDATE oec_courses SET semester=?, course_type=?, course_code=?,"
        " title=? WHERE id=?", (sem, typ, code, title, course_id))

    # ---- cascade: Main Entry Sheet records auto-update --------------------
    old_slot = _slot_for_course(old["semester"], old["course_type"])
    new_slot = _slot_for_course(sem, typ)
    now = datetime.now().isoformat(timespec="seconds")
    touched = 0
    if old_slot and old["course_code"]:
        rows = db.execute(
            "SELECT id FROM oec_students WHERE %s_code = ?" % old_slot,
            (old["course_code"],)).fetchall()
        for r in rows:
            if new_slot and new_slot != old_slot:
                # course moved to another slot → move the assignment
                db.execute(
                    "UPDATE oec_students SET %(old)s_code='',"
                    " %(old)s_title='', %(new)s_code=?, %(new)s_title=?,"
                    " updated_at=? WHERE id=?"
                    % {"old": old_slot, "new": new_slot},
                    (code, title, now, r["id"]))
            else:
                # same slot (or MJ→catalogue-only) → refresh code + title
                db.execute(
                    "UPDATE oec_students SET %(s)s_code=?, %(s)s_title=?,"
                    " updated_at=? WHERE id=?" % {"s": old_slot},
                    (code, title, now, r["id"]))
            touched += 1
    db.commit()
    flash("Course #%d updated%s." %
          (course_id,
           " — %d student record(s) on the Main Entry Sheet auto-updated"
           % touched if touched else ""), "success")
    return redirect(url_for("oec_courses"))


@app.route("/oec/courses/delete/<int:course_id>", methods=["POST"])
@admin_required
def oec_course_delete(course_id):
    db = get_db()
    db.execute("DELETE FROM oec_courses WHERE id=?", (course_id,))
    db.commit()
    flash("Course #%d deleted (student records already saved keep their "
          "text — nothing is lost)." % course_id, "success")
    return redirect(url_for("oec_courses"))


# -------------------------------------------------------------------- run
@app.route("/healthz")
def healthz():
    """Liveness/readiness probe for Render/Railway/Docker health checks."""
    try:
        db = get_db()
        db.execute("SELECT 1")
        return jsonify(status="ok", db="ok"), 200
    except Exception as exc:                                       # noqa
        return jsonify(status="error", detail=str(exc)), 500


if __name__ == "__main__":
    app.run(host=os.environ.get("HOST", "0.0.0.0"), port=PORT,
            debug=False, threaded=True)
